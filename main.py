from fastapi import FastAPI, HTTPException, Request, Form, File, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse, PlainTextResponse
from python_freeipa import Client
import os
import urllib3
from dotenv import load_dotenv
from pydantic import BaseModel, EmailStr, Field
from typing import Optional, List
import uuid
from datetime import datetime, timedelta
import json
import logging
import openpyxl
from io import BytesIO
import re
import subprocess

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('ipa-api.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

YOPASS_URL = os.getenv("YOPASS_URL")
YOPASS = os.getenv("YOPASS")
IPA_HOST = os.getenv("IPA_HOST")
SESSION_EXPIRATION_MINUTES = 60

class UserCreate(BaseModel):
    first_name: str
    last_name: str
    email: EmailStr
    title: str = None
    phone: str = None
    groups: list[str] = []  # Опционально: список групп для добавления пользователя


urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

app = FastAPI(
    title="FreeIPA API",
    description="API для управления пользователями FreeIPA",
    version="1.0.0"
)

# Хранилище сессий (в продакшене используйте Redis или базу)
user_sessions = {}
# Словарь для хранения клиентов FreeIPA по сессии
ipa_clients = {}


def resolve_username(client, identifier: str) -> str:
    """
    Преобразует identifier (username или email) в username.
    
    Args:
        client: FreeIPA клиент
        identifier: username или email
        
    Returns:
        username пользователя
        
    Raises:
        ValueError: если пользователь не найден
    """
    # Если это не email - возвращаем как есть
    if "@" not in identifier:
        return identifier
    
    # Ищем по email
    search_result = client._request(
        "user_find",
        args=[],
        params={"mail": identifier.lower()}
    )
    
    # Проверяем что нашли
    if search_result.get('count', 0) == 0:
        raise ValueError(f"Пользователь с email '{identifier}' не найден")
    
    users_list = search_result.get('result', [])
    if not users_list:
        raise ValueError(f"Пользователь с email '{identifier}' не найден")
    
    return users_list[0]['uid'][0]

def create_yopass_link(username: str, password: str):
    secret_data = f"{username}\n{password}"

    link = subprocess.run(
        [YOPASS, "--api", YOPASS_URL, "--url", YOPASS_URL, "--expiration=1w", "--one-time=true"],
        input=secret_data,
        capture_output=True,
        text=True
    )

    yopass_link = link.stdout.strip()
    return yopass_link


def cleanup_session(session_id: str):
    """Удаляет сессию и связанный FreeIPA клиент"""
    if session_id in user_sessions:
        del user_sessions[session_id]
    if session_id in ipa_clients:
        del ipa_clients[session_id]


def create_freeipa_client(host: str = None):
    """Создаёт клиент FreeIPA без авторизации"""
    host = host or IPA_HOST
    if not host:
        raise Exception("Не задан IPA_HOST в .env файле")
    
    return Client(host=host, verify_ssl=False)


def authenticate_user(username: str, password: str) -> Client:
    """Аутентификация пользователя в FreeIPA"""
    try:
        client = create_freeipa_client()
        client.login(username, password)
        return client
    except Exception as e:
        raise HTTPException(
            status_code=401,
            detail=f"Ошибка аутентификации: {str(e)}"
        )


def get_user_client(request: Request) -> Client:
    """Получает клиент FreeIPA для текущего пользователя из сессии"""
    session_id = request.cookies.get("ipa_session")
    
    if not session_id:
        raise HTTPException(status_code=401, detail="Не авторизован")
    
    if session_id not in user_sessions:
        raise HTTPException(status_code=401, detail="Сессия истекла")
    
    session_data = user_sessions[session_id]

    # Проверяем срок действия сессии
    if datetime.now() > session_data["expires"]:
        cleanup_session(session_id)
        raise HTTPException(status_code=401, detail="Сессия истекла")
    
    if session_id not in ipa_clients:
        raise HTTPException(status_code=401, detail="Ошибка сессии")
    
    return ipa_clients[session_id]



def transliterate(name):

    dictionary = {
    'а': 'a',
    'б': 'b',
    'в': 'v',
    'г': 'g',
    'д': 'd',
    'е': 'e',
    'ё': 'e',
    'ж': 'zh',
    'з': 'z',
    'и': 'i',
    'й': 'y',
    'к': 'k',
    'л': 'l',
    'м': 'm',
    'н': 'n',
    'о': 'o',
    'п': 'p',
    'р': 'r',
    'с': 's',
    'т': 't',
    'у': 'u',
    'ф': 'f',
    'х': 'kh',
    'ц': 'ts',
    'ч': 'ch',
    'ш': 'sh',
    'щ': 'sch',
    'ъ': '',
    'ы': 'y',
    'ь': '',
    'э': 'e',
    'ю': 'yu',
    'я': 'ya',
    'А': 'A',
    'Б': 'B',
    'В': 'V',
    'Г': 'G',
    'Д': 'D',
    'Е': 'E',
    'Ё': 'E',
    'Ж': 'Zh',
    'З': 'Z',
    'И': 'I',
    'Й': 'Y',
    'К': 'K',
    'Л': 'L',
    'М': 'M',
    'Н': 'N',
    'О': 'O',
    'П': 'P',
    'Р': 'R',
    'С': 'S',
    'Т': 'T',
    'У': 'U',
    'Ф': 'F',
    'Х': 'Kh',
    'Ц': 'Ts',
    'Ч': 'Ch',
    'Ш': 'Sh',
    'Щ': 'Sch',
    'Ъ': '',
    'Ы': 'Y',
    'Ь': '',
    'Э': 'E',
    'Ю': 'Yu',
    'Я': 'Ya',
    ',': '',
    '?': '',
    ' ': ' ',
    '~': '',
    '!': '',
    '@': '',
    '#': '',
    '$': '',
    '%': '',
    '^': '',
    '&': '',
    '*': '',
    '(': '',
    ')': '',
    '-': '',
    '=': '',
    '+': '',
    ':': '',
    ';': '',
    '<': '',
    '>': '',
    '\'': '',
    '"': '',
    '\\': '',
    '/': '',
    '№': '',
    '[': '',
    ']': '',
    '{': '',
    '}': '',
    'ґ': '',
    'ї': '',
    'є': '',
    'Ґ': 'g',
    'Ї': 'i',
    'Є': 'e',
    '—': ''
}

    for key in dictionary:
        name = name.replace(key, dictionary[key])
    return name

def is_valid_email(email: str) -> bool:
    """Проверка валидности email"""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

def parse_excel_row(row):
    """Парсит строку Excel"""
    return {
        "fio": str(row[0]).strip() if row[0] else "",
        "email": str(row[1]).strip() if len(row) > 1 and row[1] else "",
        "phone": str(row[2]).strip() if len(row) > 2 and row[2] else None,
        "title": str(row[3]).strip() if len(row) > 3 and row[3] else None,
        "groups_str": str(row[4]).strip() if len(row) > 4 and row[4] else ""
    }

def parse_fio(fio: str):
    """Парсит ФИО и генерирует username. Возвращает (last_name, first_name, username) или None"""
    fio_parts = fio.split()
    if len(fio_parts) < 2:
        return None

    last_name = fio_parts[0]
    first_name = fio_parts[1]
    last_name_en = transliterate(last_name).lower()
    first_name_en = transliterate(first_name).lower()
    username = f"{first_name_en}.{last_name_en}"

    return last_name, first_name, username

def parse_groups(groups_str: str):
    """Парсит строку групп через запятую"""
    if not groups_str:
        return []
    return [g.strip() for g in groups_str.split(',') if g.strip()]

@app.post("/login")
async def login(request: Request, username: str = Form(...), password: str = Form(...)):
    """Аутентификация пользователя в FreeIPA"""
    try:
        logger.info(f"Login attempt: {username}")
        # Аутентифицируем пользователя в FreeIPA
        client = authenticate_user(username, password)

        # Создаём сессию
        session_id = str(uuid.uuid4())
        expires = datetime.now() + timedelta(minutes=SESSION_EXPIRATION_MINUTES)
        
        # Сохраняем данные сессии
        user_sessions[session_id] = {
            "username": username,
            "created": datetime.now(),
            "expires": expires
        }
        
        # Сохраняем клиент FreeIPA
        ipa_clients[session_id] = client
        
        # Создаём ответ с кукой
        response = JSONResponse(
            content={
                "status": "ok",
                "user": username,
                "session_id": session_id
            }
        )
        
        response.set_cookie(
            key="ipa_session",
            value=session_id,
            httponly=True,
            max_age=SESSION_EXPIRATION_MINUTES * 60,
            path="/"
        )

        logger.info(f"Login successful: {username}")
        return response

    except Exception as e:
        logger.warning(f"Login failed: {username} - {str(e)}")
        raise HTTPException(status_code=401, detail=f"Ошибка авторизации: {str(e)}")



@app.post("/logout")
async def logout(request: Request):
    """Выход из системы"""
    session_id = request.cookies.get("ipa_session")

    if session_id:
        cleanup_session(session_id)

    response = JSONResponse(content={"status": "logged out"})
    response.delete_cookie("ipa_session", path="/")
    return response


@app.get("/api/v1/users/{username}")
def get_user(username: str, request: Request):
    """
    Получение информации о пользователе
    Аналог в FreeIPA: ipa user-show --all username
    """

    try:
        # Получаем клиент из сессии
        client = get_user_client(request)
        
        # Получаем информацию о пользователе
        user = client._request("user_show", args=[username], params={"all": True})
        return user
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка получения пользователя: {str(e)}"
        )


@app.post("/api/v1/utils/text-to-json")
async def text_to_json(users_text: str):
    """
    Отдаем:
    ivan@test.com
    petr@test.com
    anna@test.com

    Получаем:
    ["ivan@test.com", "petr@test.com", "anna@test.com"]
    """
    identifiers = [
        email.strip() 
        for line in users_text.split('\n') 
        if line.strip() 
        for email in line.split()
    ]
    return identifiers


@app.post("/api/v1/users/{username}/delete")
def delete_user(username: str, request: Request):
    """
    Обычное удаление пользователя в FreeIPA

    Удаляет одного пользователя

    Использовать с умом т.к безвозвратно удаляет пользователя в FreeIPA
    """
    try:
        session_id = request.cookies.get("ipa_session")
        admin = user_sessions.get(session_id, {}).get("username", "unknown")
        logger.warning(f"USER_DELETE: {username} by {admin}")

        client = get_user_client(request)
        result = client._request("user_del", args=[username], params={})
        
        logger.info(f"USER_DELETE SUCCESS: {username}")
        return {
            "username": username,
            "message": f"Пользователь {username} успешно удалён",
            "status": "deleted"
        }

    except Exception as e:
        logger.error(f"USER_DELETE FAILED: {username} - {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка удаления пользователя: {str(e)}"
        )

@app.post("/api/v1/users/{username}/disable")
def disable_user(username: str, request: Request):
    """
    Обычное выключение пользователя в FreeIPA

    Выключение одного пользователя
    """
    try:
        client = get_user_client(request)
        result = client._request("user_disable", args=[username], params={})

        return {
            "username": username,
            "message": f"Пользователь {username} успешно отключен",
            "status": "disable"
        }

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка отключения пользователя: {str(e)}"
        )

@app.post("/api/v1/users/{username}/enable")
def enable_user(username: str, request: Request):
    """
    Активация юзера в FreeIPA

    Активация одного пользователя
    """
    try:
        client = get_user_client(request)
        result = client._request("user_enable", args=[username], params={})

        return {
            "username": username,
            "message": f"Пользователь {username} успешно включен",
            "status": "enable"
        }

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка активации пользователя: {str(e)}"
        )

@app.post("/api/v1/users/{username}/reset-password")
def reset_password(username: str, request: Request):
    """
    Сброс пароля пользователя в FreeIPA

    Пароль генерируется автоматически и возвращается далее для работы

    Также возвращается дата до какого числа будет действовать пароль `"expiration": result['result'].get('mail', [None])[0]`

    Вводим только логин УЗ
    """
    try:
        session_id = request.cookies.get("ipa_session")
        admin = user_sessions.get(session_id, {}).get("username", "unknown")
        logger.warning(f"PASSWORD_RESET: {username} by {admin}")

        client = get_user_client(request)
        result = client._request("user_mod", args=[username], params={"random": True})

        password = result['result']['randompassword']

        yopass_link = create_yopass_link(username,password)

        response = {
            "username": username,
            "password": password,
            "yopass_link": yopass_link,
            "expiration": result['result'].get('krbpasswordexpiration', [None])[0],
            "message": f"Пароль пользователя {username} успешно сброшен"
        }

        logger.info(f"PASSWORD_RESET SUCCESS: {username}")
        return response

    except Exception as e:
        logger.error(f"PASSWORD_RESET FAILED: {username} - {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка сброса пароля: {str(e)}"
        )


@app.post("/api/v1/creat-users")
def create_user(user: UserCreate, request: Request):
    """
    Создает пользователя c передачей цельного JSON файла

    Можно указать список групп для автоматического добавления:

    {
        "first_name": "Иван",
        "last_name": "Иванов",
        "email": "ivan@test.com",
        "groups": ["admins", "developers"]
    }

    ВАЖНО: Если указаны группы и ни одна не добавилась - пользователь будет удален и вернется ошибка.
    """
    try:
        username = f"{user.first_name.lower()}.{user.last_name.lower()}"
        full_name = f"{user.first_name} {user.last_name}"

        session_id = request.cookies.get("ipa_session")
        admin = user_sessions.get(session_id, {}).get("username", "unknown")
        logger.info(f"USER_CREATE: {username} ({user.email}) by {admin}")

        client = get_user_client(request)

        # Создаём пользователя тут ipa user_add
        result = client._request(
            "user_add",
            args=[username],
            params={
                "givenname": user.first_name,
                "sn": user.last_name,
                "cn": full_name,
                "mail": user.email,
                "title": user.title,
                "telephonenumber": user.phone,
                "random": True,
            }
        )

        password = result['result']['randompassword']

        # Тут вызываю новый метод group_add_member т.к в user_add нет такого функционала
        added_groups = []
        failed_groups = []

        for group in user.groups:
            try:
                client._request(
                    "group_add_member",
                    args=[group],
                    params={"user": username}
                )
                added_groups.append(group)
            except Exception as e:
                failed_groups.append({"group": group, "error": str(e)})

        response = {
            "username": username,
            "password": password,
            "email": user.email, # Сразу добавил чтобы выводил почту чтобы передавать её дальше для генерации ссылки в passoc и отправки
            "message": f"Пользователь {username} создан"
        }

        # Проверяем результат добавления в группы
        if user.groups:
            # Если группы были указаны, но НИ ОДНА не добавилась - откатываю создание пользователя
            if len(added_groups) == 0:
                try:
                    client._request("user_del", args=[username], params={})
                except:
                    pass  # Игнорирую ошибки удаления

                raise HTTPException(
                    status_code=500,
                    detail=f"Пользователь создан, но не удалось добавить ни в одну группу. Пользователь удален. Ошибки: {failed_groups}"
                )

            # Если хотя бы одна группа добавилась - возвращаем успех с информацией
            response["groups"] = {
                "added": added_groups,
                "failed": failed_groups
            }

        logger.info(f"USER_CREATE SUCCESS: {username} with groups {added_groups}")
        return response

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"USER_CREATE FAILED: {username} - {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/v1/users/create-form")
def create_user_form(
    request: Request,
    first_name: str = Form(...),
    last_name: str = Form(...),
    email: str = Form(...),
    title: Optional[str] = Form(None),
    phone: Optional[str] = Form(None),
    groups: Optional[str] = Form(None)
):
    """
    Это аналог ручки /api/v1/creat-users только уже не в формате JSON чтобы ручками его не создавать
    

    ВАЖНО: Если указаны группы и ни одна не добавилась - пользователь будет удален и вернется ошибка.
    """
    try:
        username = f"{first_name.lower()}.{last_name.lower()}" # Считаю это идеальной УЗ, правки будут внесены если надо maybe
        full_name = f"{first_name} {last_name}"

        client = get_user_client(request)

        result = client._request(
            "user_add",
            args=[username],
            params={
                "givenname": first_name,
                "sn": last_name,
                "cn": full_name,
                "mail": email,
                "title": title,
                "telephonenumber": phone,
                "random": True,
            }
        )

        password = result['result']['randompassword']

        added_groups = []
        failed_groups = []

        # Парсим строку с группами (разделенные запятыми)
        groups_list = []
        if groups and groups.strip(): # Проверяю что groups не null
            groups_list = [g.strip() for g in groups.split(',') if g.strip()]   # тут split убирает пробелы т.к 100 процентов будут ошибки и split чтобы разбивать если несколько групп то есть ["admins", "dev", "ops"]

        for group in groups_list:
            try:
                client._request(
                    "group_add_member",
                    args=[group],
                    params={"user": username}
                )
                added_groups.append(group)
            except Exception as e:
                failed_groups.append({"group": group, "error": str(e)})

        response = {
            "username": username,
            "password": password,
            "email": email,
            "message": f"Пользователь {username} создан"
        }

        if groups_list:
            if len(added_groups) == 0:
                try:
                    client._request("user_del", args=[username], params={})
                except:
                    pass
                raise HTTPException(
                    status_code=500,
                    detail=f"Пользователь создан, но не удалось добавить ни в одну группу. Пользователь удален. Ошибки: {failed_groups}"
                )
            response["groups"] = {
                "added": added_groups,
                "failed": failed_groups
            }

        return response

    except HTTPException:
        raise  # перебрасываем HTTPException дальше 
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


#-------------------Массовые операции-----------------
# Это массовые ручки для работы с большим количеством пользователей
# Ручки с bulk для массовых операций 

@app.post("/api/v1/users/bulk-delete")
def bulk_delete_users(identifiers: list[str], request: Request):
    """
    Массовое удаление пользователей
    
    Принимает username или email

    ["ivan.ivanov", "petr@test.com", "petya.petrov"]
    """
    results = {"success": [], "failed": []}
    client = get_user_client(request)

    for identifier in identifiers:
        try:
            # Находим username (по email или напрямую)
            username = resolve_username(client, identifier)
            
            # Удаляем пользователя
            client._request("user_del", args=[username], params={})
            
            # Добавляем в успешные
            results["success"].append({
                "identifier": identifier,
                "username": username
            })
            
        except ValueError as e:
            # Пользователь не найден
            results["failed"].append({
                "identifier": identifier,
                "error": str(e)
            })
        except Exception as e:
            # Любая другая ошибка (FreeIPA, сеть и т.д.)
            results["failed"].append({
                "identifier": identifier,
                "error": f"Ошибка удаления: {str(e)}"
            })

    return results


@app.post("/api/v1/users/bulk-disable")
def bulk_disable_users(identifiers: list[str], request: Request):
    """
    Массовое удаление пользователей
    
    Принимает username или email

    ["ivan.ivanov", "petr@test.com", "petya.petrov"]
    """
    results = {"success": [], "failed": []}
    client = get_user_client(request)

    for identifier in identifiers:
        try:
            # Находим username (по email или напрямую)
            username = resolve_username(client, identifier)
            
            # Удаляем пользователя
            client._request("user_disable", args=[username], params={})
            
            # Добавляем в успешные
            results["success"].append({
                "identifier": identifier,
                "username": username
            })
            
        except ValueError as e:
            # Пользователь не найден
            results["failed"].append({
                "identifier": identifier,
                "error": str(e)
            })
        except Exception as e:
            # Любая другая ошибка (FreeIPA, сеть и т.д.)
            results["failed"].append({
                "identifier": identifier,
                "error": f"Ошибка отключения: {str(e)}"
            })

    return results


@app.post("/api/v1/users/bulk-enable")
def bulk_enable_users(identifiers: list[str], request: Request):
    """
    Массовое включение пользователей
    
    Принимает username или email

    ["ivan.ivanov", "petr@test.com", "petya.petrov"]
    """
    results = {"success": [], "failed": []}
    client = get_user_client(request)

    for identifier in identifiers:
        try:
            # Находим username (по email или напрямую)
            username = resolve_username(client, identifier)
            
            # Удаляем пользователя
            client._request("user_enable", args=[username], params={})
            
            # Добавляем в успешные
            results["success"].append({
                "identifier": identifier,
                "username": username
            })
            
        except ValueError as e:
            # Пользователь не найден
            results["failed"].append({
                "identifier": identifier,
                "error": str(e)
            })
        except Exception as e:
            # Любая другая ошибка (FreeIPA, сеть и т.д.)
            results["failed"].append({
                "identifier": identifier,
                "error": f"Ошибка включения: {str(e)}"
            })

    return results


@app.post("/api/v1/users/bulk-reset-password")
def bulk_reset_password(identifiers: list[str], request: Request):
    """
    Массовый сброс паролей пользователей

    Можно передавать username или email - API сам определит:
    ["ivan.ivanov", "petr@test.com", "elena.sidorova"]
    """
    results = {
        "success": [],
        "failed": []
    }

    client = get_user_client(request)

    for identifier in identifiers:
        try:
            # Находим username (вся логика внутри функции)
            username = resolve_username(client, identifier)  # ← ВОТ И ВСЁ!

            # Сбрасываем пароль
            reset_result = client._request(
                "user_mod",
                args=[username],
                params={"random": True}
            )

            password = reset_result['result']['randompassword']

            results["success"].append({
                "identifier": identifier,
                "username": username,
                "password": password
            })

        except Exception as e:
            results["failed"].append({
                "identifier": identifier,
                "error": str(e)
            })

    return results


@app.get("/api/v1/users/search-by-email/{email}")
def search_user_by_email(email: str, request: Request):
    """
    Поиск пользователя по email (для отладки)

    Показывает что именно возвращает FreeIPA при поиске
    """
    try:
        client = get_user_client(request)

        # Пробуем разные варианты поиска
        results = {}

        # 1. Точное совпадение
        try:
            search_exact = client._request(
                "user_find",
                args=[],
                params={"mail": email}
            )
            results["exact_match"] = {
                "query": email,
                "type": str(type(search_exact)),
                "is_dict": isinstance(search_exact, dict),
                "is_list": isinstance(search_exact, list),
                "data": search_exact
            }
        except Exception as e:
            results["exact_match"] = {"error": str(e)}

        # 2. Lowercase
        try:
            search_lower = client._request(
                "user_find",
                args=[],
                params={"mail": email.lower()}
            )
            results["lowercase"] = {
                "query": email.lower(),
                "type": str(type(search_lower)),
                "is_dict": isinstance(search_lower, dict),
                "is_list": isinstance(search_lower, list),
                "data": search_lower
            }
        except Exception as e:
            results["lowercase"] = {"error": str(e)}

        # 3. Поиск по username (если это вдруг username)
        try:
            user_show = client._request("user_show", args=[email.split('@')[0]], params={})
            results["by_username"] = user_show.get('result', {})
        except Exception as e:
            results["by_username"] = {"error": str(e)}

        return results

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/v1/report/full-usersgroups-info")
def fullusersgroupsinfo(request: Request):
    """
    Получение информации о всех пользователях и его группах
    """
    try:
        client = get_user_client(request)
        result = client._request("user_find", args=[], params={"all":True})

        user_data = []
        csv_lines = ["username,email,groups"]

        for user in result['result']:
            username = user['uid'][0]
            email = user.get('mail', [None])[0]
            groups = user.get('memberof_group', [])

            user_info = {
                "username": username,
                "email": email,
                "groups": groups
            }
            user_data.append(user_info)

            email_str = email or ''
            groups_str = ';'.join(groups)
            csv_lines.append(f"{username},{email_str},{groups_str}")

        csv_content = '\n'.join(csv_lines)
        return StreamingResponse(
            iter([csv_content]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=users_groups_report.csv"}
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка: {str(e)}"
        )

@app.get("/api/v1/report/full-info")
def full_info(request: Request):
    """
    Получение информации о всех пользователях и его группах
    """
    try:
        client = get_user_client(request)
        result = client._request("user_find", args=[], params={"all":True})

        return result

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка: {str(e)}"
        )

@app.post("/api/v1/users/validate-excel")
async def validate_excel(request: Request, file: UploadFile = File(...)):
    """
    Валидация Excel файла перед массовым созданием пользователей

    Проверяет:
    - Формат email
    - Конфликты username (уже существует в FreeIPA)
    - Конфликты email (уже существует в FreeIPA)
    - Дубликаты email внутри файла
    - Существование групп
    - Корректность ФИО (минимум 2 слова)

    Возвращает детальный отчёт БЕЗ создания пользователей
    """
    try:
        # Проверяем авторизацию
        client = get_user_client(request)

        session_id = request.cookies.get("ipa_session")
        admin = user_sessions.get(session_id, {}).get("username", "unknown")
        logger.info(f"VALIDATE_EXCEL: Started by {admin}")

        # Читаем Excel файл
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents))
        sheet = workbook.active

        conflicts = []
        warnings = []
        would_create = 0
        emails_in_file = {}  # Для отслеживания дубликатов внутри файла

        # Получаем список всех существующих пользователей из FreeIPA
        existing_users = client._request("user_find", args=[], params={"all": True})
        existing_usernames = {u['uid'][0] for u in existing_users['result']}

        # Собираем существующие email (безопасно)
        existing_emails = set()
        for u in existing_users['result']:
            mail = u.get('mail')
            if mail and isinstance(mail, list) and len(mail) > 0 and mail[0]:
                existing_emails.add(mail[0].lower())

        # Кэш для проверки групп (чтобы не проверять одну группу несколько раз)
        groups_cache = {}

        # Проходим по строкам
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Пропускаем пустые строки
            if not row or not row[0]:
                continue

            try:
                # Извлекаем данные
                fio = str(row[0]).strip() if row[0] else ""
                email = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                phone = str(row[2]).strip() if len(row) > 2 and row[2] else None
                title = str(row[3]).strip() if len(row) > 3 and row[3] else None
                groups_str = str(row[4]).strip() if len(row) > 4 and row[4] else ""

                # Проверка 1: ФИО заполнено
                if not fio:
                    conflicts.append({
                        "row": row_num,
                        "error": "ФИО не заполнено"
                    })
                    continue

                # Проверка 2: Email заполнен
                if not email:
                    conflicts.append({
                        "row": row_num,
                        "fio": fio,
                        "error": "Email не заполнен"
                    })
                    continue

                # Проверка 3: Email валидный
                if not is_valid_email(email):
                    conflicts.append({
                        "row": row_num,
                        "fio": fio,
                        "error": f"Невалидный email: {email}"
                    })
                    continue

                # Проверка 4: Email дубликат внутри файла
                email_lower = email.lower()
                if email_lower in emails_in_file:
                    conflicts.append({
                        "row": row_num,
                        "fio": fio,
                        "error": f"Дубликат email {email} (уже в строке {emails_in_file[email_lower]})"
                    })
                    continue
                emails_in_file[email_lower] = row_num

                # Парсим ФИО
                fio_parts = fio.split()
                if len(fio_parts) < 2:
                    conflicts.append({
                        "row": row_num,
                        "fio": fio,
                        "error": "ФИО должно содержать минимум Фамилию и Имя"
                    })
                    continue

                last_name = fio_parts[0]
                first_name = fio_parts[1]

                # Генерируем username
                last_name_en = transliterate(last_name).lower()
                first_name_en = transliterate(first_name).lower()
                username = f"{first_name_en}.{last_name_en}"

                # Проверка 5 и 6: Собираем все конфликты для этой строки
                row_errors = []

                # Проверка 5: Username уже существует в FreeIPA
                if username in existing_usernames:
                    row_errors.append(f"Username '{username}' уже существует в FreeIPA")

                # Проверка 6: Email уже существует в FreeIPA
                if email_lower in existing_emails:
                    row_errors.append(f"Email '{email}' уже существует в FreeIPA")

                # Если есть конфликты - добавляем и пропускаем строку
                if row_errors:
                    conflicts.append({
                        "row": row_num,
                        "fio": fio,
                        "username": username,
                        "email": email,
                        "error": "; ".join(row_errors)
                    })
                    continue

                # Проверка 7: Существование групп
                if groups_str:
                    groups_list = [g.strip() for g in groups_str.split(',') if g.strip()]
                    non_existing_groups = []

                    for group in groups_list:
                        # Проверяем кэш
                        if group not in groups_cache:
                            try:
                                client._request("group_show", args=[group])
                                groups_cache[group] = True
                            except:
                                groups_cache[group] = False

                        if not groups_cache[group]:
                            non_existing_groups.append(group)

                    if non_existing_groups:
                        conflicts.append({
                            "row": row_num,
                            "fio": fio,
                            "username": username,
                            "error": f"Группы не существуют: {', '.join(non_existing_groups)}"
                        })
                        continue

                # Предупреждения (не блокируют создание)
                if not phone:
                    warnings.append({
                        "row": row_num,
                        "fio": fio,
                        "username": username,
                        "message": "Телефон не заполнен"
                    })

                if not title:
                    warnings.append({
                        "row": row_num,
                        "fio": fio,
                        "username": username,
                        "message": "Должность не заполнена"
                    })

                # Если всё ок - считаем как валидного
                would_create += 1

            except Exception as e:
                conflicts.append({
                    "row": row_num,
                    "fio": fio if 'fio' in locals() else "unknown",
                    "error": f"Неожиданная ошибка: {str(e)}"
                })

        # Формируем результат
        total_rows = sheet.max_row - 1  # Минус заголовок
        valid = len(conflicts) == 0

        result = {
            "valid": valid,
            "total_rows": total_rows,
            "would_create": would_create,
            "conflicts_count": len(conflicts),
            "warnings_count": len(warnings),
            "conflicts": conflicts,
            "warnings": warnings
        }

        logger.info(f"VALIDATE_EXCEL: Completed by {admin} - Valid: {valid}, Would create: {would_create}, Conflicts: {len(conflicts)}")
        return result

    except Exception as e:
        logger.error(f"VALIDATE_EXCEL: Critical error - {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка валидации Excel файла: {str(e)}"
        )

@app.post("/api/v1/users/bulk-create-from-excel")
async def bulk_create_from_excel(request: Request, file: UploadFile = File(...)):
    """
    Парсинг excel и создание пользователя
    """
    try:
        # Сначала проверяем авторизацию (до чтения файла!)
        client = get_user_client(request)

        session_id = request.cookies.get("ipa_session")
        admin = user_sessions.get(session_id, {}).get("username", "unknown")
        logger.info(f"BULK_CREATE_EXCEL: Started by {admin}")

        # Читаем Excel файл (только если авторизован)
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents)) # превращаем биты в читаемый файл
        sheet = workbook.active # Активный листы

        # Проверяем доступность Yopass ДО начала создания пользователей
        try:
            test_link = create_yopass_link("test", "test123")
            logger.info(f"BULK_CREATE_EXCEL: Yopass check OK - {test_link}")
        except Exception as e:
            logger.error(f"BULK_CREATE_EXCEL: Yopass unavailable - {str(e)}")
            raise HTTPException(
                status_code=503,
                detail=f"Yopass недоступен: {str(e)}. Создание пользователей отменено."
            )

        results = {"success": [], "failed": []}

        # Проходим по строкам (пропускаем первую - заголовки)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Пропускаем пустые строки
            if not row or not row[0]:
                continue

            try:
                # Парсим строку Excel
                data = parse_excel_row(row)
                fio = data["fio"]
                email = data["email"]

                # Валидация обязательных полей
                if not fio:
                    results["failed"].append({"row": row_num, "error": "ФИО не заполнено"})
                    continue

                if not email:
                    results["failed"].append({"row": row_num, "fio": fio, "error": "Email не заполнен"})
                    continue

                if not is_valid_email(email):
                    results["failed"].append({"row": row_num, "fio": fio, "error": f"Невалидный email: {email}"})
                    continue

                # Парсим ФИО
                fio_parsed = parse_fio(fio)
                if not fio_parsed:
                    results["failed"].append({"row": row_num, "fio": fio, "error": "ФИО должно содержать минимум Фамилию и Имя"})
                    continue

                last_name, first_name, username = fio_parsed

                # Собираем все ошибки валидации для этой строки
                row_errors = []

                # Проверка: Username уже существует в FreeIPA
                try:
                    client._request("user_show", args=[username])
                    # Если не упало - значит пользователь существует
                    row_errors.append(f"Username '{username}' уже существует в FreeIPA")
                except:
                    # Пользователь не найден - можно создавать
                    pass

                # Проверка: Email уже существует в FreeIPA
                try:
                    # Ищем пользователей с таким email
                    email_check = client._request("user_find", args=[], params={"mail": email})
                    if email_check['result']:
                        # Найден пользователь с таким email
                        existing_username = email_check['result'][0]['uid'][0]
                        row_errors.append(f"Email '{email}' уже используется пользователем {existing_username}")
                except:
                    # Ошибка поиска - игнорируем и продолжаем
                    pass

                # Парсим группы
                groups_list = parse_groups(data["groups_str"])

                # Проверка: Существование всех групп
                if groups_list:
                    non_existing_groups = []
                    for group in groups_list:
                        try:
                            client._request("group_show", args=[group])
                        except:
                            non_existing_groups.append(group)

                    if non_existing_groups:
                        row_errors.append(f"Группы не существуют: {', '.join(non_existing_groups)}")

                # Если есть любые ошибки валидации - не создаём пользователя
                if row_errors:
                    results["failed"].append({
                        "row": row_num,
                        "fio": fio,
                        "username": username,
                        "email": email,
                        "error": "; ".join(row_errors)
                    })
                    continue

                # Создаём пользователя в FreeIPA
                result = client._request(
                    "user_add",
                    args=[username],
                    params={
                        "givenname": first_name,
                        "sn": last_name,
                        "cn": fio,
                        "mail": email,
                        "title": data["title"],
                        "telephonenumber": data["phone"],
                        "random": True,
                    }
                )

                password = result['result']['randompassword']

                yopass_link = create_yopass_link(username, password)

                # Добавляем в группы
                added_groups = []
                failed_groups = []

                for group in groups_list:
                    try:
                        client._request(
                            "group_add_member",
                            args=[group],
                            params={"user": username}
                        )
                        added_groups.append(group)
                    except Exception as e:
                        failed_groups.append({"group": group, "error": str(e)})

                # Если валидация прошла успесното добавляем сюда
                success_entry = {
                    "row": row_num,
                    "fio": fio,
                    "username": username,
                    "email": email,
                    "password": password,
                    "yopass_link": yopass_link
                }

                if groups_list:
                    success_entry["groups"] = {
                        "added": added_groups,
                        "failed": failed_groups
                    }

                results["success"].append(success_entry)
                logger.info(f"BULK_CREATE_EXCEL: Created {username} from row {row_num}")

            except Exception as e:
                results["failed"].append({
                    "row": row_num,
                    "fio": fio if 'fio' in locals() else "unknown",
                    "error": str(e)
                })
                logger.error(f"BULK_CREATE_EXCEL: Failed row {row_num} - {str(e)}")

        logger.info(f"BULK_CREATE_EXCEL: Completed by {admin} - Success: {len(results['success'])}, Failed: {len(results['failed'])}")

        return results

    except Exception as e:
        logger.error(f"BULK_CREATE_EXCEL: Critical error - {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка обработки Excel файла: {str(e)}"
        )




@app.post("/api/v1/yopass/echo", response_class=PlainTextResponse)
def generate_yopass_link(data: str = Form()):
    """Генерирует Yopass ссылку через форму"""
    try:
        result = subprocess.run(
            [
                YOPASS,
                "--api", YOPASS_URL,
                "--url", YOPASS_URL,
                "--expiration=1w",
                "--one-time=true"
            ],
            input=data,
            capture_output=True,
            text=True,
            check=True
        )
        logger.info(f"Ссылка успешно сгенерирована")
        return result.stdout.strip()

        
    except subprocess.CalledProcessError as e:
        logger.error(f"Ошибка - {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка {e.stderr}"
        )
    except FileNotFoundError:
        logger.error(f"Yopass binary не найден. Проверьте путь - {str(e)}")
        raise HTTPException(
            status_code=500,
            detail="Yopass binary не найден. Проверьте путь"
        )

if __name__ == "__main__":
    import uvicorn
    
    if not IPA_HOST:
        print("⚠️  Внимание: IPA_HOST не задан в .env файле!")
        print("Создайте файл .env с содержимым:")
        print("IPA_HOST=",IPA_HOST)
    
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=8080,
        reload=True
    )